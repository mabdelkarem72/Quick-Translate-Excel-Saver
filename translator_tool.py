import os
import time
import pyperclip
import keyboard
from deep_translator import GoogleTranslator
from openpyxl import Workbook, load_workbook
from plyer import notification

FILE_NAME = "words.xlsx"
HOTKEY = "ctrl+shift+k" 
last_execution_time = 0

def save_word():
    global last_execution_time
    current_time = time.time()
    
    if current_time - last_execution_time < 1.5:
        return
    
    word = pyperclip.paste().strip()
    
    if not word or len(word) > 50:
        return
    
    try:
        translation = GoogleTranslator(source='auto', target='ar').translate(word)
        
        if not os.path.exists(FILE_NAME):
            wb = Workbook()
            ws = wb.active
            ws.append(["الكلمة", "الترجمة"])
            wb.save(FILE_NAME)

        try:
            wb = load_workbook(FILE_NAME)
            ws = wb.active
            ws.append([word, translation])
            wb.save(FILE_NAME)
            
            notification.notify(
                title="✅ تم الحفظ في الإكسيل",
                message=f"الكلمة: {word}\nالترجمة: {translation}",
                app_name="مترجم محمود",
                timeout=3 
            
        except PermissionError:
            notification.notify(
                title="❌ خطأ: الملف مفتوح",
                message="يرجى إغلاق ملف words.xlsx ليتمكن البرنامج من الحفظ",
                timeout=5
            )
            
        last_execution_time = current_time
        
    except Exception as e:
        print(f"حدث خطأ: {e}")

if __name__ == "__main__":
    print(f"البرنامج شغال.. انسخ أي كلمة ودوس {HOTKEY}")
    keyboard.add_hotkey(HOTKEY, save_word)
    keyboard.wait()